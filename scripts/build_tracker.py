#!/usr/bin/env python3
"""
LinkedIn Hand-Raiser Engagement Tracker Builder
Generates production-grade Excel workbook with formulas, formatting, and sample data pre-built.
Includes video tracking for Runway AI and Loom videos.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from datetime import datetime, timedelta
import os

def create_tracker_workbook():
    """Create the complete LinkedIn engagement tracker workbook"""

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create the three main sheets
    create_post_performance_sheet(wb)
    create_dashboard_sheet(wb)
    create_response_tracker_sheet(wb)

    # Save to tracking folder
    output_path = os.path.expanduser('~/tk_projects/linkedin-hand-raiser-system/tracking/linkedin_engagement_tracker.xlsx')
    wb.save(output_path)
    print(f"✅ Tracker created successfully: {output_path}")
    return output_path

def create_post_performance_sheet(wb):
    """TAB 1: Post Performance Tracker with Video Tracking"""
    ws = wb.create_sheet("Post Performance Tracker", 0)

    # Define headers (NEW: Video columns F-I added)
    headers = [
        'Post ID', 'Date Posted', 'Vertical', 'Pain Point', 'Post Type',
        'Video Type', 'Video Style', 'Video URL', 'Video Cost',  # NEW VIDEO COLUMNS
        'Impressions', 'Likes', 'Comments', 'Shares', 'DMs Received',
        'Videos Sent', 'Demos Booked', 'Notes',
        'Total Engagement', 'Engagement %', 'Conversion %',
        'Video→Demo %', 'Cost/Engage', 'Tier', 'Status'
    ]

    # Write headers (Row 1)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data (6 posts with video tracking)
    # Columns: Post ID, Date, Vertical, Pain Point, Post Type,
    #          Video Type, Video Style, Video URL, Video Cost,
    #          Impressions, Likes, Comments, Shares, DMs, Videos Sent, Demos, Notes
    sample_data = [
        ['EC-001', datetime(2025, 1, 13), 'Electrical', 'Material Cost Overruns', 'Hand-Raiser',
         'Runway', 'Pain Point Illustration', 'https://app.runwayml.com/video/ec001', 2.50,
         1247, 63, 8, 2, 5, 5, 2, 'Runway video: copper price spike animation - high engagement'],
        ['EC-002', datetime(2025, 1, 15), 'Electrical', 'Labor Burden Blindness', 'Hand-Raiser',
         'None', 'N/A', '', 0,
         892, 34, 3, 1, 2, 2, 0, 'Text-only baseline test for comparison'],
        ['HV-001', datetime(2025, 1, 15), 'HVAC', 'Labor Burden Blindness', 'Hand-Raiser',
         'Loom', 'Personal Demo', 'https://loom.com/share/hv001demo', 0,
         1583, 87, 12, 4, 8, 7, 3, 'Personal Loom explanation - strong resonance on "80% done"'],
        ['PL-001', datetime(2025, 1, 17), 'Plumbing', 'Admin Time Waste', 'Hand-Raiser',
         'Runway', 'Text Overlay', 'https://app.runwayml.com/video/pl001', 1.75,
         1129, 56, 6, 2, 4, 4, 1, 'Runway text animation: "3 HOURS A DAY" - office manager empathy'],
        ['HV-002', datetime(2025, 1, 20), 'HVAC', 'Material Cost Overruns', 'Hand-Raiser',
         'Runway', 'Abstract Motion', 'https://app.runwayml.com/video/hv002', 2.00,
         743, 28, 2, 0, 1, 1, 0, 'Abstract motion graphics - lower fit for HVAC audience'],
        ['EC-003', datetime(2025, 1, 22), 'Electrical', 'Cash Flow Delays', 'Educational',
         'None', 'N/A', '', 0,
         1034, 41, 5, 1, 3, 3, 1, 'Testing new pain point angle without video']
    ]

    # Write sample data (Rows 2-7)
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Add formulas for calculated columns
    # NEW COLUMN MAPPING (shifted by 4):
    # R = Total Engagement, S = Engagement %, T = Conversion %
    # U = Video→Demo %, V = Cost/Engage, W = Tier, X = Status
    for row in range(2, 8):  # Rows 2-7
        # Column R: Total Engagement (SUM of Likes + Comments + Shares) - was N, now R
        # Likes=K, Comments=L, Shares=M
        ws[f'R{row}'] = f'=SUM(K{row}:M{row})'

        # Column S: Engagement % (Total Engagement / Impressions * 100) - was O, now S
        # Impressions=J
        ws[f'S{row}'] = f'=IF(J{row}>0,R{row}/J{row}*100,0)'
        ws[f'S{row}'].number_format = '0.0"%"'

        # Column T: Conversion % ((Comments + DMs) / Impressions * 100) - was P, now T
        # Comments=L, DMs=N
        ws[f'T{row}'] = f'=IF(J{row}>0,(L{row}+N{row})/J{row}*100,0)'
        ws[f'T{row}'].number_format = '0.0"%"'

        # Column U: Video→Demo % (Demos / Videos Sent * 100) - was Q, now U
        # Videos Sent=O, Demos=P
        ws[f'U{row}'] = f'=IF(O{row}>0,P{row}/O{row}*100,0)'
        ws[f'U{row}'].number_format = '0.0"%"'

        # Column V: Cost/Engage (Video Cost / Total Engagement) - was R, now V
        # Video Cost=I, Total Engagement=R
        ws[f'V{row}'] = f'=IF(R{row}>0,I{row}/R{row},0)'
        ws[f'V{row}'].number_format = '$0.00'

        # Column W: Performance Tier (based on Engagement %) - was S, now W
        ws[f'W{row}'] = f'=IF(S{row}>=5,"🟢 High",IF(S{row}>=2,"🟡 Medium","🔴 Low"))'

        # Column X: Status (pipeline stage indicator) - was T, now X
        # Demos=P, Videos Sent=O, DMs=N
        ws[f'X{row}'] = f'=IF(P{row}>0,"✅ Closed",IF(O{row}>0,"📹 Video Sent",IF(N{row}>0,"💬 Engaged","👀 Posted")))'

    # Apply conditional formatting
    # Engagement % heatmap (Column S, was O)
    ws.conditional_formatting.add(
        'S2:S100',
        ColorScaleRule(
            start_type='num', start_value=0, start_color='EA4335',
            mid_type='num', mid_value=5, mid_color='FBBC04',
            end_type='num', end_value=10, end_color='34A853'
        )
    )

    # Performance Tier highlighting (Column W, was S)
    ws.conditional_formatting.add(
        'W2:W100',
        CellIsRule(operator='containsText', formula=['"High"'], fill=PatternFill(start_color='D4EDDA', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'W2:W100',
        CellIsRule(operator='containsText', formula=['"Medium"'], fill=PatternFill(start_color='FFF3CD', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'W2:W100',
        CellIsRule(operator='containsText', formula=['"Low"'], fill=PatternFill(start_color='F8D7DA', fill_type='solid'))
    )

    # Demo booked highlight (Column P, was L)
    ws.conditional_formatting.add(
        'P2:P100',
        CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill(start_color='34A853', fill_type='solid'), font=Font(color='FFFFFF', bold=True))
    )

    # Video Type highlighting
    ws.conditional_formatting.add(
        'F2:F100',
        CellIsRule(operator='containsText', formula=['"Runway"'], fill=PatternFill(start_color='E8D5E8', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'F2:F100',
        CellIsRule(operator='containsText', formula=['"Loom"'], fill=PatternFill(start_color='D5E8D5', fill_type='solid'))
    )

    # Adjust column widths (updated for new columns)
    column_widths = {
        'A': 10, 'B': 12, 'C': 12, 'D': 20, 'E': 12,
        'F': 12, 'G': 22, 'H': 35, 'I': 10,  # NEW VIDEO COLUMNS
        'J': 12, 'K': 8, 'L': 10, 'M': 8, 'N': 12,
        'O': 12, 'P': 12, 'Q': 35, 'R': 12, 'S': 12,
        'T': 12, 'U': 12, 'V': 12, 'W': 12, 'X': 15
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Add data validation dropdowns
    from openpyxl.worksheet.datavalidation import DataValidation

    # Vertical dropdown (Column C)
    dv_vertical = DataValidation(type="list", formula1='"Electrical,HVAC,Plumbing,Energy,Multi-Vertical"', allow_blank=False)
    ws.add_data_validation(dv_vertical)
    dv_vertical.add('C2:C100')

    # Pain Point dropdown (Column D)
    dv_pain = DataValidation(type="list", formula1='"Material Cost Overruns,Labor Burden Blindness,Admin Time Waste,Cash Flow Delays,Job Profitability Opacity"', allow_blank=False)
    ws.add_data_validation(dv_pain)
    dv_pain.add('D2:D100')

    # Post Type dropdown (Column E)
    dv_type = DataValidation(type="list", formula1='"Hand-Raiser,Educational,Case Study,Poll/Survey,Testimonial"', allow_blank=False)
    ws.add_data_validation(dv_type)
    dv_type.add('E2:E100')

    # NEW: Video Type dropdown (Column F)
    dv_video_type = DataValidation(type="list", formula1='"None,Loom,Runway"', allow_blank=False)
    ws.add_data_validation(dv_video_type)
    dv_video_type.add('F2:F100')

    # NEW: Video Style dropdown (Column G)
    dv_video_style = DataValidation(type="list", formula1='"N/A,Personal Demo,Abstract Motion,Pain Point Illustration,Text Overlay,Before-After"', allow_blank=False)
    ws.add_data_validation(dv_video_style)
    dv_video_style.add('G2:G100')

def create_dashboard_sheet(wb):
    """TAB 2: Vertical Performance Dashboard with Video Analytics"""
    ws = wb.create_sheet("Dashboard", 1)

    # Section 1: Vertical Performance Summary
    ws['A1'] = 'VERTICAL PERFORMANCE SUMMARY'
    ws['A1'].font = Font(bold=True, size=14)

    headers = ['Vertical', 'Total Posts', 'Avg Engage%', 'Total DMs', 'Videos Sent', 'Demos Booked']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    verticals = ['Electrical', 'HVAC', 'Plumbing', 'TOTAL']
    for row_num, vertical in enumerate(verticals, 3):
        ws.cell(row=row_num, column=1, value=vertical)

    # Formulas for Electrical (Row 3) - UPDATED COLUMN REFS
    ws['B3'] = '=COUNTIF(\'Post Performance Tracker\'!C:C,"Electrical")'
    ws['C3'] = '=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!C:C,"Electrical",\'Post Performance Tracker\'!S:S),0)'
    ws['C3'].number_format = '0.0"%"'
    ws['D3'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Electrical",\'Post Performance Tracker\'!N:N)'
    ws['E3'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Electrical",\'Post Performance Tracker\'!O:O)'
    ws['F3'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Electrical",\'Post Performance Tracker\'!P:P)'

    # Formulas for HVAC (Row 4)
    ws['B4'] = '=COUNTIF(\'Post Performance Tracker\'!C:C,"HVAC")'
    ws['C4'] = '=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!C:C,"HVAC",\'Post Performance Tracker\'!S:S),0)'
    ws['C4'].number_format = '0.0"%"'
    ws['D4'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"HVAC",\'Post Performance Tracker\'!N:N)'
    ws['E4'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"HVAC",\'Post Performance Tracker\'!O:O)'
    ws['F4'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"HVAC",\'Post Performance Tracker\'!P:P)'

    # Formulas for Plumbing (Row 5)
    ws['B5'] = '=COUNTIF(\'Post Performance Tracker\'!C:C,"Plumbing")'
    ws['C5'] = '=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!C:C,"Plumbing",\'Post Performance Tracker\'!S:S),0)'
    ws['C5'].number_format = '0.0"%"'
    ws['D5'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Plumbing",\'Post Performance Tracker\'!N:N)'
    ws['E5'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Plumbing",\'Post Performance Tracker\'!O:O)'
    ws['F5'] = '=SUMIF(\'Post Performance Tracker\'!C:C,"Plumbing",\'Post Performance Tracker\'!P:P)'

    # TOTAL row formulas (Row 6)
    ws['B6'] = '=SUM(B3:B5)'
    ws['C6'] = '=AVERAGE(C3:C5)'
    ws['C6'].number_format = '0.0"%"'
    ws['D6'] = '=SUM(D3:D5)'
    ws['E6'] = '=SUM(E3:E5)'
    ws['F6'] = '=SUM(F3:F5)'

    # Bold TOTAL row
    for col in range(1, 7):
        ws.cell(row=6, column=col).font = Font(bold=True)

    # ========================================
    # NEW SECTION: Video Performance Analysis
    # ========================================
    ws['A9'] = 'VIDEO PERFORMANCE ANALYSIS'
    ws['A9'].font = Font(bold=True, size=14)

    video_headers = ['Video Type', 'Posts', 'Avg Engage%', 'Total DMs', 'Demos', 'Lift vs None']
    for col_num, header in enumerate(video_headers, 1):
        cell = ws.cell(row=10, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    video_types = ['None', 'Loom', 'Runway']
    for row_num, vtype in enumerate(video_types, 11):
        ws.cell(row=row_num, column=1, value=vtype)
        ws[f'B{row_num}'] = f'=COUNTIF(\'Post Performance Tracker\'!F:F,"{vtype}")'
        ws[f'C{row_num}'] = f'=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!F:F,"{vtype}",\'Post Performance Tracker\'!S:S),0)'
        ws[f'C{row_num}'].number_format = '0.0"%"'
        ws[f'D{row_num}'] = f'=SUMIF(\'Post Performance Tracker\'!F:F,"{vtype}",\'Post Performance Tracker\'!N:N)'
        ws[f'E{row_num}'] = f'=SUMIF(\'Post Performance Tracker\'!F:F,"{vtype}",\'Post Performance Tracker\'!P:P)'
        # Lift vs None (compare to row 11 which is "None")
        if vtype == 'None':
            ws[f'F{row_num}'] = '-'
        else:
            ws[f'F{row_num}'] = f'=IF(C11>0,C{row_num}/C11-1,0)'
            ws[f'F{row_num}'].number_format = '+0.0%;-0.0%'

    # Video Style Performance (Runway Only)
    ws['A16'] = 'RUNWAY VIDEO STYLE PERFORMANCE'
    ws['A16'].font = Font(bold=True, size=14)

    style_headers = ['Video Style', 'Posts', 'Avg Engage%', 'DMs', 'Demos']
    for col_num, header in enumerate(style_headers, 1):
        cell = ws.cell(row=17, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    styles = ['Abstract Motion', 'Pain Point Illustration', 'Text Overlay', 'Before-After']
    for row_num, style in enumerate(styles, 18):
        ws.cell(row=row_num, column=1, value=style)
        ws[f'B{row_num}'] = f'=COUNTIF(\'Post Performance Tracker\'!G:G,"{style}")'
        ws[f'C{row_num}'] = f'=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!G:G,"{style}",\'Post Performance Tracker\'!S:S),0)'
        ws[f'C{row_num}'].number_format = '0.0"%"'
        ws[f'D{row_num}'] = f'=SUMIF(\'Post Performance Tracker\'!G:G,"{style}",\'Post Performance Tracker\'!N:N)'
        ws[f'E{row_num}'] = f'=SUMIF(\'Post Performance Tracker\'!G:G,"{style}",\'Post Performance Tracker\'!P:P)'

    # ========================================
    # Section: Pain Point Performance
    # ========================================
    ws['A24'] = 'PAIN POINT PERFORMANCE'
    ws['A24'].font = Font(bold=True, size=14)

    pain_headers = ['Pain Point', 'Posts Count', 'Avg Engage%', 'Demos']
    for col_num, header in enumerate(pain_headers, 1):
        cell = ws.cell(row=25, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    pain_points = ['Material Cost Overruns', 'Labor Burden Blindness', 'Admin Time Waste', 'Cash Flow Delays', 'Job Profitability Opacity']
    for row_num, pain in enumerate(pain_points, 26):
        ws.cell(row=row_num, column=1, value=pain)
        ws[f'B{row_num}'] = f'=COUNTIF(\'Post Performance Tracker\'!D:D,"{pain}")'
        ws[f'C{row_num}'] = f'=IFERROR(AVERAGEIF(\'Post Performance Tracker\'!D:D,"{pain}",\'Post Performance Tracker\'!S:S),0)'
        ws[f'C{row_num}'].number_format = '0.0"%"'
        ws[f'D{row_num}'] = f'=SUMIF(\'Post Performance Tracker\'!D:D,"{pain}",\'Post Performance Tracker\'!P:P)'

    # ========================================
    # Section: Conversion Funnel
    # ========================================
    ws['A33'] = 'CONVERSION FUNNEL'
    ws['A33'].font = Font(bold=True, size=14)

    funnel_stages = [
        ('Total Impressions', '=SUM(\'Post Performance Tracker\'!J:J)'),
        ('Total Engagement', '=SUM(\'Post Performance Tracker\'!R:R)'),
        ('DMs Received', '=SUM(\'Post Performance Tracker\'!N:N)'),
        ('Videos Sent', '=SUM(\'Post Performance Tracker\'!O:O)'),
        ('Demos Booked', '=SUM(\'Post Performance Tracker\'!P:P)')
    ]

    ws['A34'] = 'Stage'
    ws['B34'] = 'Count'
    ws['A34'].font = Font(bold=True)
    ws['B34'].font = Font(bold=True)

    for row_num, (stage, formula) in enumerate(funnel_stages, 35):
        ws.cell(row=row_num, column=1, value=stage)
        ws.cell(row=row_num, column=2, value=formula)

    # ========================================
    # Section: Video ROI Summary
    # ========================================
    ws['A42'] = 'VIDEO ROI SUMMARY'
    ws['A42'].font = Font(bold=True, size=14)

    ws['A43'] = 'Total Video Cost'
    ws['B43'] = '=SUM(\'Post Performance Tracker\'!I:I)'
    ws['B43'].number_format = '$0.00'

    ws['A44'] = 'Cost per Demo (Video Posts)'
    ws['B44'] = '=IF(SUM(\'Post Performance Tracker\'!P:P)>0,B43/SUMIF(\'Post Performance Tracker\'!F:F,"<>None",\'Post Performance Tracker\'!P:P),0)'
    ws['B44'].number_format = '$0.00'

    ws['A45'] = 'Video Engagement Premium'
    ws['B45'] = '=IF(C11>0,(C12+C13)/2/C11-1,0)'
    ws['B45'].number_format = '+0.0%;-0.0%'

    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_response_tracker_sheet(wb):
    """TAB 3: Engagement Response Tracker"""
    ws = wb.create_sheet("Response Tracker", 2)

    # Define headers
    headers = [
        'Prospect', 'Vertical', 'Post ID', 'Engage Type', 'Date Engaged',
        'Video Sent', 'Demo Date', 'Status', 'Days→Video', 'Days→Demo',
        'Speed Grade', 'Stage'
    ]

    # Write headers (Row 1)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data
    sample_prospects = [
        ['John Smith', 'Electrical', 'EC-001', 'Comment', datetime(2025, 1, 13),
         datetime(2025, 1, 13), datetime(2025, 1, 18), 'Demo Booked'],
        ['Mike Johnson', 'HVAC', 'HV-001', 'DM', datetime(2025, 1, 15),
         datetime(2025, 1, 15), datetime(2025, 1, 20), 'Demo Booked'],
        ['Sarah Williams', 'Plumbing', 'PL-001', 'Comment', datetime(2025, 1, 17),
         datetime(2025, 1, 18), None, 'Video Sent'],
        ['Dave Martinez', 'Electrical', 'EC-001', 'DM', datetime(2025, 1, 13),
         None, None, 'Engaged']
    ]

    # Write sample data
    for row_num, row_data in enumerate(sample_prospects, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

    # Add formulas for calculated columns
    for row in range(2, 6):  # Rows 2-5
        # Column I: Days→Video (Video Sent - Date Engaged)
        ws[f'I{row}'] = f'=IF(F{row}="","",F{row}-E{row})'

        # Column J: Days→Demo (Demo Date - Date Engaged)
        ws[f'J{row}'] = f'=IF(G{row}="","",G{row}-E{row})'

        # Column K: Speed Grade (based on Days→Video)
        ws[f'K{row}'] = f'=IF(I{row}="","",IF(I{row}<=1,"🟢 Fast",IF(I{row}<=3,"🟡 OK","🔴 Slow")))'

        # Column L: Pipeline Stage
        ws[f'L{row}'] = f'=IF(H{row}="Demo Booked","Stage 3: Demo",IF(F{row}<>"","Stage 2: Video Sent",IF(E{row}<>"","Stage 1: Engaged","")))'

    # Apply conditional formatting
    # Overdue video responses (Column I > 2 days)
    ws.conditional_formatting.add(
        'I2:I100',
        CellIsRule(operator='greaterThan', formula=['2'], fill=PatternFill(start_color='EA4335', fill_type='solid'), font=Font(color='FFFFFF', bold=True))
    )

    # Pipeline stage coloring (Column L)
    ws.conditional_formatting.add(
        'L2:L100',
        CellIsRule(operator='containsText', formula=['"Stage 3"'], fill=PatternFill(start_color='34A853', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'L2:L100',
        CellIsRule(operator='containsText', formula=['"Stage 2"'], fill=PatternFill(start_color='FBBC04', fill_type='solid'))
    )
    ws.conditional_formatting.add(
        'L2:L100',
        CellIsRule(operator='containsText', formula=['"Stage 1"'], fill=PatternFill(start_color='4A90E2', fill_type='solid'))
    )

    # Adjust column widths
    column_widths = {
        'A': 18, 'B': 12, 'C': 10, 'D': 12, 'E': 14,
        'F': 14, 'G': 14, 'H': 15, 'I': 12, 'J': 12,
        'K': 15, 'L': 18
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

if __name__ == "__main__":
    print("🚀 Building LinkedIn Hand-Raiser Engagement Tracker...")
    print("   📊 Adding video tracking columns (Runway, Loom)")
    print("   📈 Adding video performance dashboard")
    create_tracker_workbook()
    print("\n✅ Complete! Open the file to start tracking.")
